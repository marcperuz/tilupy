# -*- coding: utf-8 -*-

from typing import Callable

import math as math
import numpy as np
import matplotlib
import matplotlib.pyplot as plt

import tilupy.read
import tilupy.notations as notations
from tilupy.analytic_sol import Coussot_shape


class Benchmark:
    """Benchmark of simulation results.
    
    This class groups together all the methods for processing and analyzing 
    the results of various simulation models, allowing, among other things, 
    the comparison of results between models.
    
    Attributes
    ----------
        _loaded_results : dict[tilupy.read.Results]
            Dictionnary of :class:`tilupy.read.Results` object for each model specified in 
            :meth:`load_numerical_result`.
        _models_tim : dict[list]
            Dictionnary of recorded time steps for each model specified in 
            :meth:`load_numerical_result`.
        _simu_z : numpy.ndarray
            Topographic elevation for the simulation loaded.
    """
    def __init__(self):
        self._loaded_results = {}
        self._models_tim = {}
        self._simu_z = None
        
    
    def load_numerical_result(self, 
                              model: str, 
                              **kwargs,
                              ) -> None:
        """Load numerical simulation results using :func:`tilupy.read.get_results` for a given model and
        store the :class:`tilupy.read.Results` object in :attr:`_loaded_results`.

        Parameters
        ----------
        model : str
            Name of the model to load. Must be one of the :data:`tilupy.read.ALLOWED_MODELS`.
        **kwargs
            Keyword arguments passed to the result reader for the specific model.

        Raises
        ------
        ValueError
            If size of stored :attr:`_simu_z` is different with new loaded ones.
        UserWarning
            If stored :attr:`_simu_z` are different with new loaded ones.
        ValueError
            If the provided model is not in the allowed list of :data:`tilupy.read.ALLOWED_MODELS`.
        """
        if model in tilupy.read.ALLOWED_MODELS:
            if model not in self._loaded_results:
                self._loaded_results[model] = tilupy.read.get_results(model, **kwargs)
                self._models_tim[model] = self._loaded_results[model].tim
                
            if self._simu_z is None:
                self._simu_z = self._loaded_results[model].z
            
            if self._simu_z.shape != self._loaded_results[model].z.shape:
                raise ValueError("Simulations loaded doesn't have same size.")
            
            try :
                if not np.allclose(self._simu_z, self._loaded_results[model].z, rtol=0.1, atol=0.1):
                    mean_error = np.mean(self._simu_z - self._loaded_results[model].z)
                    self._simu_z = self._loaded_results[model].z
                    raise UserWarning(f"Stored elevation values different from loaded ones; average error: {mean_error}")
            except UserWarning as w:
                print(f"[WARNING] {w}")
                
        else:
            raise ValueError(f" -> No correct model selected, choose between:\n       {tilupy.read.ALLOWED_MODELS}")
    
       
    def compute_analytic_solution(self,
                                  output: str,
                                  solution: Callable,
                                  model: str,
                                  T: float | list[float], 
                                  **kwargs
                                  ) -> None:
        """
        Compute the analytic solution for the given time steps using the provided model.

        Parameters
        ----------
        output : str
            Wanted output for the analytical solution. Can be : "h" or "u".
        solution : Callable
            Callable object representing the analytic solution model (model from :class:`tilupy.analytic_sol.Depth_result`)
        model : str
            Wanted model to compute the analytical solution from. Must be loaded first with 
            :meth:`load_numerical_result`.
        T : float | list[float]
            Time or list of times at which to compute the analytic solution.
        **kwargs
            Keyword arguments passed to the analytic solution for the specific model.
        
        Raises
        ------
        ValueError
            If no solution found.
        """
        solution = solution(**kwargs)
        
        if isinstance(T, float) or isinstance(T, int):
            T = [T]
        
        if output not in ['h', 'u']:
            raise ValueError(" -> Available output: 'h', 'u'.")
        
        if output == 'h':
            solution.compute_h(self._loaded_results[model].x, T)
            
            if solution.h is not None:
                return tilupy.read.TemporalResults1D(name=output,
                                                     d=solution.h[:].T,
                                                     t=T,
                                                     coords=self._loaded_results[model].x,
                                                     coords_name='x')
            else:
                raise ValueError("No analytic solution for fluid height.")
            
        if output == 'u':
            solution.compute_u(self._loaded_results[model].x, T)
            
            if solution.u is not None:
                return tilupy.read.TemporalResults1D(name=output,
                                                     d=solution.u[:].T,
                                                     t=T,
                                                     coords=self._loaded_results[model].x,
                                                     coords_name='x')
            else:
                raise ValueError("No analytic solution for fluid velocity.")


    def show_output(self,
                    output: str,
                    model: str,
                    time_steps: float | list[float] = None,
                    show_plot: bool = True,
                    **plot_kwargs,
                    ) -> matplotlib.axes._axes.Axes:
        """Plot output for a given model using :meth:`tilupy.read.Results.plot`.

        Parameters
        ----------
        output : str
            Wanted output, need to be in :data:`tilupy.read.DATA_NAMES`.
        model : str
           Wanted model to show the data field. Must be loaded with :meth:`load_numerical_result`
        time_steps : float | list[float]
            Value or list of time steps display output. 
            If None displays the output for all recorded time steps in the model's result. 
            By default None.
        show_plot : bool, optional
            If True, show the plot, by default True.

        Returns
        -------
        matplotlib.axes._axes.Axes
            The created plot.

        Raises
        ------
        ValueError
            If model is not loaded.
        """
        if model not in self._loaded_results.keys():
            raise ValueError(" -> First load model using load_numerical_result.")
        
        plot_kwargs = {} if plot_kwargs is None else plot_kwargs
        
        ax = self._loaded_results[model].plot(output=output,
                                              time_steps=time_steps,
                                              display_plot=show_plot,
                                              **plot_kwargs)
        if show_plot:
            plt.show()
        
        return ax


    def show_output_profile(self,
                            output: str,
                            model: str,
                            extraction_method: str = "axis",
                            extraction_params: dict = None,
                            time_steps: float | list[float] = None,
                            show_plot: bool = True,
                            **plot_kwargs
                            )-> matplotlib.axes._axes.Axes:
        """Plot profile from 2D temporal or static output for a given model using 
        :meth:`tilupy.read.Results.plot_profile`.

        Parameters
        ----------
        output : str
            Wanted output, need to be in :data:`tilupy.read.TEMPORAL_DATA_2D` or :data:`tilupy.read.STATIC_DATA_2D`
        model : str
            Wanted model to show the output profile. Must be loaded with :meth:`load_numerical_result`
        extraction_method : str, optional
            Wanted profile extraction method. See :func:`tilupy.utils.get_profile`. By default "axis".
        extraction_params : dict, optional
            Profile extraction parameters. See :func:`tilupy.utils.get_profile`. Be default None.
        time_steps : float | list[float], optional
            Value or list of time steps required to extract and display profiles. 
            If None displays the profiles for all recorded time steps in the model's result. 
            Only available for models. By default None.
        show_plot : bool, optional
            If True, show the plot, by default True.

        Returns
        -------
        matplotlib.axes._axes.Axes
            The created plot.

        Raises
        ------
        ValueError
            If model is not loaded.
        """
        if model not in self._loaded_results.keys():
            raise ValueError(" -> First load model using load_numerical_result")

        plot_kwargs = {} if plot_kwargs is None else plot_kwargs
        
        ax = self._loaded_results[model].plot_profile(output=output,
                                                      extraction_method = extraction_method,
                                                      extraction_params = extraction_params,
                                                      time_steps=time_steps,
                                                      display_plot=show_plot,
                                                      **plot_kwargs)
        if show_plot:
            plt.show()
            
        return ax


    def show_comparison_temporal1D(self,
                                   output: str,
                                   models: list[str],
                                   profile_extraction_args: dict = None,
                                   analytic_solution: dict = None,
                                   time_steps: float | list[float] = None,
                                   axes: matplotlib.axes._axes.Axes = None,
                                   rows_nb: int = None,
                                   cols_nb: int = None,
                                   figsize: tuple[float] = None,
                                   colors: list = None,
                                   linestyles: list[str] = None,
                                   plot_kwargs: dict = None,
                                   as_kwargs: dict = None,
                                   show_plot: bool = True,
                                   ) -> matplotlib.axes._axes.Axes:
        """Plot multiple temporal 1D data in the same graph.

        Parameters
        ----------
        output : str
            Wanted output, need to be in :data:`tilupy.read.TEMPORAL_DATA_1D`. If using a :data:`tilupy.read.TEMPORAL_DATA_2D`,
            must use :data:`profile_extraction_args`.
        models : list[str]
            Wanted models to show the temporal 1D data. Must be loaded with :meth:`load_numerical_result`.
        profile_extraction_args : dict, optional
            Arguments for profile extraction. See :func:`tilupy.utils.get_profile`. By default None.
        analytic_solution : dict, optional
            Arguments for plotting an analytic solution. See :meth:`compute_analytic_solution`. 
            By default None.
        time_steps : float | list[float], optional
            Value or list of time steps required to extract and display profiles. 
            If None displays the profiles for all recorded time steps in the model's result. 
            Only available for models. By default None.
        axes : matplotlib.axes._axes.Axes, optional
            Existing matplotlib axe, by default None.
        rows_nb : int, optional
            Number of rows when plotting multiple time steps. If None automatically computed. 
            By default None.
        cols_nb : int, optional
            Number of colums when plotting multiple time steps. If None use 3 columns. 
            By default None.
        figsize : tuple[float], optional
            Size of the plotted figure (Height, Width), by default None.
        colors : list, optional
            List of colors for each model plotted. If None use "black". By default None.
        linestyles : list[str], optional
            List of linestyles for each model plotted. If None use regular straight line. By default None.
        plot_kwargs : dict, optional
            Additional argument for model's line, by default None.
        as_kwargs : dict, optional
            Additional argument for analytical solution's line, by default None.
        show_plot : bool, optional
            If True, show the plot, by default True.

        Returns
        -------
        matplotlib.axes._axes.Axes
            The created plot.

        Raises
        ------
        ValueError
            If models is not loaded.
        """
        for model in models:
            if model not in self._loaded_results.keys():
                raise ValueError(" -> First load model using load_numerical_result.")
                
        plot_kwargs = {} if plot_kwargs is None else plot_kwargs
        as_kwargs = {} if as_kwargs is None else as_kwargs
        profile_extraction_args = {} if profile_extraction_args is None else profile_extraction_args
        if "extraction_method" not in profile_extraction_args:
            profile_extraction_args["extraction_method"] = "axis"
        
        profile_models = {}
        for model in models:
            if output in tilupy.read.TEMPORAL_DATA_2D:
                prof, data = self._loaded_results[model].get_profile(output=output,
                                                                     **profile_extraction_args)
                profile_models[model] = [prof, data]
            elif output in tilupy.read.TEMPORAL_DATA_1D:
                prof = self._loaded_results[model].get_output(output)
                profile_models[model] = [prof, None]
                
        if time_steps is None:
            time_steps = self._models_tim[list(self._loaded_results.keys())[0]]
        
        if analytic_solution is not None:
            as_profile = self.compute_analytic_solution(output=output,
                                                        model=list(self._loaded_results.keys())[0],
                                                        T=time_steps,
                                                        **analytic_solution)
        
        if isinstance(time_steps, float) or isinstance(time_steps, int):
            time_steps = np.array([time_steps])
        elif isinstance(time_steps, list):
            time_steps = np.array(time_steps)
        
        if isinstance(profile_models[model][0], tilupy.read.TemporalResults):
            for model in models:
                profile_models[model][0] = profile_models[model][0].extract_from_time_step(time_steps)
        
        if axes is None:
            if cols_nb is None:
                cols_nb = len(time_steps) if len(time_steps) < 3 else 3
            
            if rows_nb is None:
                rows_nb = len(time_steps) // cols_nb
                if len(time_steps) % cols_nb != 0:
                    rows_nb += 1
                    
            fig, axes = plt.subplots(nrows=rows_nb, 
                                     ncols=cols_nb, 
                                     figsize=figsize, 
                                     layout="constrained", 
                                     sharex=True, 
                                     sharey=True)
            if isinstance(axes, plt.Axes):
                axes = np.array([axes])
            else:
                axes = axes.flatten()
                
        for T in range(len(time_steps)) :
            # Plot models
            for i in range(len(models)):
                if colors is not None and len(colors) > i:
                    color = colors[i]
                else:
                    color = "black"
                if linestyles is not None and len(linestyles) > i:
                    linestyle = linestyles[i]
                else:
                    linestyle = None
                
                if "alpha" not in plot_kwargs:
                    plot_kwargs["alpha"] = 0.8
                if "linewidth" not in plot_kwargs:
                    plot_kwargs["linewidth"] = 1.5
                
                axes[T].plot(profile_models[models[i]][0].coords, profile_models[models[i]][0].d[:, T], color=color, linestyle=linestyle, label=models[i], **plot_kwargs)
            
            # Plot analytic solution
            if analytic_solution is not None:
                if "color" not in as_kwargs:
                    as_kwargs["color"] = "red"
                if "alpha" not in as_kwargs:
                    as_kwargs["alpha"] = 0.9
                if "linewidth" not in as_kwargs:
                    as_kwargs["linewidth"] = 1
                
                axes[T].plot(as_profile.coords, as_profile.d[:, T], label=f"{str(analytic_solution['solution']).split('.')[-1][:-2]}", **as_kwargs)
            
            # Formatting fig                
            axes[T].set_xlim(left=min(profile_models[models[0]][0].coords), right=max(profile_models[models[0]][0].coords))
            axes[T].grid(True, alpha=0.3)
            
            if len(time_steps) == 1:
                if profile_extraction_args["extraction_method"] == "axis" and output in tilupy.read.TEMPORAL_DATA_2D:
                    inv_axis = ""
                    if profile_models[models[0]][0].coords_name == 'x':
                        inv_axis = "Y"
                    else:
                        inv_axis = "X"
                    axes[T].set_title(f"{inv_axis}={profile_models[models[0]][1]}m  |  t={time_steps[T]}s")
                else:
                    axes[T].set_title(f"t={time_steps[T]}s")
            else:
                axes[T].set_title(f"t={time_steps[T]}s", loc="left")
                if profile_extraction_args["extraction_method"] == "axis" and output in tilupy.read.TEMPORAL_DATA_2D:
                    inv_axis = ""
                    if profile_models[models[0]][0].coords_name == 'x':
                        inv_axis = "Y"
                    else:
                        inv_axis = "X"
                    axes[T].set_title(f"{inv_axis}={profile_models[models[0]][1]}m", loc="right")
                
            axes[T].set_xlabel(notations.get_label(profile_models[models[0]][0].coords_name))
            axes[T].set_ylabel(notations.get_label(output))
            axes[T].legend(loc='upper right')

        for i in range(len(time_steps), len(axes)):
            fig.delaxes(axes[i])    
        
        if show_plot:
            plt.show()
        
        return axes

    
    def get_avrg_result(self, 
                        output: str
                        ) -> tilupy.read.TemporalResults | tilupy.read.StaticResults:
        """Get average result computed with all loaded model (:meth:`load_numerical_result`).

        Parameters
        ----------
        output : str
            Wanted average output.

        Returns
        -------
        tilupy.read.TemporalResults | tilupy.read.StaticResults
            Average result output.

        Raises
        ------
        ValueError
            If StaticResults0D output wanted.
        """
        output_list = []
        for model in self._loaded_results:
            output_list.append(self._loaded_results[model].get_output(output))
        
        if isinstance(output_list[0], tilupy.read.TemporalResults2D):
            data_list = [output_list[i].d for i in range(len(output_list))]
            time_list = [output_list[i].t for i in range(len(output_list))]
            x_list = [output_list[i].x for i in range(len(output_list))]
            y_list = [output_list[i].y for i in range(len(output_list))]
            z_list = [output_list[i].z for i in range(len(output_list))]

            mean_data = np.mean(np.stack(data_list), axis=0)
            mean_time = np.mean(np.stack(time_list), axis=0)
            mean_x = np.mean(np.stack(x_list), axis=0)
            mean_y = np.mean(np.stack(y_list), axis=0)
            mean_z = np.mean(np.stack(z_list), axis=0)
            
            mean_result = tilupy.read.TemporalResults2D(name=output_list[0].name,
                                                        d=mean_data,
                                                        t=mean_time,
                                                        x=mean_x,
                                                        y=mean_y,
                                                        z=mean_z)

            return mean_result
        
        elif isinstance(output_list[0], tilupy.read.TemporalResults1D):
            data_list = [output_list[i].d for i in range(len(output_list))]
            time_list = [output_list[i].t for i in range(len(output_list))]
            coords_list = [output_list[i].coords for i in range(len(output_list))]
            
            mean_data = np.mean(np.stack(data_list), axis=0)
            mean_time = np.mean(np.stack(time_list), axis=0)
            mean_coords = np.mean(np.stack(coords_list), axis=0)
            
            mean_result = tilupy.read.TemporalResults1D(name=output_list[0].name,
                                                        d=mean_data,
                                                        t=mean_time,
                                                        coords=mean_coords,
                                                        coords_name=output_list[0].coords_name)

            return mean_result

        elif isinstance(output_list[0], tilupy.read.TemporalResults0D):
            data_list = [output_list[i].d for i in range(len(output_list))]
            time_list = [output_list[i].t for i in range(len(output_list))]
            
            mean_data = np.mean(np.stack(data_list), axis=0)
            mean_time = np.mean(np.stack(time_list), axis=0)
            
            mean_result = tilupy.read.TemporalResults0D(name=output_list[0].name,
                                                        d=mean_data,
                                                        t=mean_time,
                                                        scalar_names=output_list[0].scalar_names)

            return mean_result

        elif isinstance(output_list[0], tilupy.read.StaticResults2D):
            data_list = [output_list[i].d for i in range(len(output_list))]
            x_list = [output_list[i].x for i in range(len(output_list))]
            y_list = [output_list[i].y for i in range(len(output_list))]
            z_list = [output_list[i].z for i in range(len(output_list))]

            mean_data = np.mean(np.stack(data_list), axis=0)
            mean_x = np.mean(np.stack(x_list), axis=0)
            mean_y = np.mean(np.stack(y_list), axis=0)
            mean_z = np.mean(np.stack(z_list), axis=0)
            
            mean_result = tilupy.read.StaticResults2D(name=output_list[0].name,
                                                      d=mean_data,
                                                      x=mean_x,
                                                      y=mean_y,
                                                      z=mean_z)

            return mean_result
        
        elif isinstance(output_list[0], tilupy.read.StaticResults1D):
            data_list = [output_list[i].d for i in range(len(output_list))]
            coords_list = [output_list[i].coords for i in range(len(output_list))]
            
            mean_data = np.mean(np.stack(data_list), axis=0)
            mean_coords = np.mean(np.stack(coords_list), axis=0)
            
            mean_result = tilupy.read.StaticResults1D(name=output_list[0].name,
                                                      d=mean_data,
                                                      coords=mean_coords,
                                                      coords_name=output_list[0].coords_name)

            return mean_result

        else:
            raise ValueError("Not available for StaticResults0D.")
        

    def compute_area(self,
                     flow_threshold: float = None
                     ) -> tuple[dict, dict]:
        """Compute area at each recorded time steps, computed with 'h', for each model loaded
        using :meth:`load_numerical_result`.

        Parameters
        ----------
        flow_threshold : float, optional
            Flow threshold to extract flow area, if None use 1% of initial maximal 
            flow height. By default None.

        Returns
        -------
        tuple[dict, dict]
            area_surf: dict
                2D area for each model: area_surf[model] = TemporalResults2D.
            area_num: dict
                Area value for each model: area_num[model] = TemporalResults0D. 
        """
        output_list = []
        for model in self._loaded_results:
            output_list.append(self._loaded_results[model].get_output("h"))
        
        height_list = [output_list[i].d for i in range(len(output_list))]

        if flow_threshold is None:
            flow_threshold = np.max(height_list[0]) * 0.01
        
        for h in height_list:
            h[h<flow_threshold] = 0
            h[h>=flow_threshold] = 1
        
        area_num = {}
        area_surf = {}
        model_name = list(self._loaded_results.keys())
        for i in range(len(height_list)):
            surface_list = []
            dx = self._loaded_results[model].dx
            dy = self._loaded_results[model].dy
            cell_surface = dx*dy

            for t in range(height_list[i].shape[2]):
                nb_cell = np.sum(height_list[i][:, :, t] == 1)
                surface_list.append(nb_cell*cell_surface)
            
            area_num[model_name[i]] = tilupy.read.TemporalResults0D(name='s',
                                                                    d=np.array(surface_list),
                                                                    t=output_list[i].t,
                                                                    scalar_names="Surface")
        
            area_surf[model_name[i]] = tilupy.read.TemporalResults2D(name='s',
                                                                     d=height_list[i],
                                                                     t=output_list[i].t,
                                                                     x=output_list[i].x,
                                                                     y=output_list[i].y,
                                                                     z=output_list[i].z)
        
        return area_surf, area_num 
    
    
    def compute_impacted_area(self,
                                flow_threshold: float = None,
                                ) -> tuple[dict, dict]:
        """Compute impacted area, computed with 'h_max', for each model loaded using 
        :meth:`load_numerical_result`.

        Parameters
        ----------
        flow_threshold : float, optional
            Flow threshold to extract flow area, if None use 1% of initial maximal 
            flow height. By default None.

        Returns
        -------
        tuple[dict, dict]
            area_surf: dict
                2D area for each model: area_surf[model] = StaticResults2D.
            area_num: dict
                Area value for each model: area_num[model] = area_val. 
        """
        output_list = []
        for model in self._loaded_results:
            output_list.append(self._loaded_results[model].get_output("h_max"))
        
        height_list = [output_list[i].d for i in range(len(output_list))]

        if flow_threshold is None:
            flow_threshold = np.max(height_list[0]) * 0.01
        
        for h in height_list:
            h[h<flow_threshold] = 0
            h[h>=flow_threshold] = 1
        
        area_num = {}
        area_surf = {}
        model_name = list(self._loaded_results.keys())
        for i in range(len(height_list)):
            dx = self._loaded_results[model].dx
            dy = self._loaded_results[model].dy
            cell_surface = dx*dy

            nb_cell = np.sum(height_list[i][:, :] == 1)
            surface = (nb_cell*cell_surface)
            
            area_num[model_name[i]] = surface
        
            area_surf[model_name[i]] = tilupy.read.StaticResults2D(name='s',
                                                                   d=height_list[i],
                                                                   x=output_list[i].x,
                                                                   y=output_list[i].y,
                                                                   z=output_list[i].z)
        
        return area_surf, area_num 
    
    
    def compute_impacted_area_rms_from_avrg(self,
                                              flow_threshold: float = None
                                              ) -> tuple[dict, np.ndarray]:
        """Compute impacted area, computed with 'h_max', RMS with average result for each model loaded
        using :meth:`load_numerical_result`.

        Parameters
        ----------
        flow_threshold : float, optional
            Flow threshold to extract flow area, if None use 1% of initial maximal 
            flow height. By default None.

        Returns
        -------
        tuple[dict, np.ndarray]
            area_rms: dict
                RMS value for each model: area_rms[model] = rms_value.
            avrg_area: numpy.ndarray
                Average area surface.
        """
        area_surf, area_num = self.compute_impacted_area(flow_threshold=flow_threshold)
        
        mean_area = []
        for model in area_surf:
            mean_area.append(np.nan_to_num(area_surf[model].d))
        
        avrg_area = np.mean(np.array(mean_area), axis=0)
                
        area_rms = {}
        for model in area_num:
            rms = (np.sqrt(np.sum((area_surf[model].d - avrg_area)**2)) /
                   np.sqrt(np.sum((avrg_area)**2)))
            area_rms[model] = rms

        return area_rms, avrg_area


    def compute_rms_from_avrg(self,
                              output: str,
                              ) -> dict:
        """Compute RMS with average result for each model loaded using :meth:`load_numerical_result`.

        Parameters
        ----------
        output : str
            Wanted output to compute the RMS.

        Returns
        -------
        dict
            RMS value for each model: output_rms[model] = rms_value.
        """
        avrg_result = self.get_avrg_result(output)
            
        extracted_data = {}
        for model in self._loaded_results:
            data = self._loaded_results[model].get_output(output)
            extracted_data[model] = data.d
        avrg_height = avrg_result.d
        
        output_rms = {}
        for model in self._loaded_results:
            rms = (np.sqrt(np.sum((extracted_data[model] - avrg_height)**2)) /
                   np.sqrt(np.sum((avrg_height)**2)))
            output_rms[model] = rms
                
        return output_rms
    
    
    def compute_dist_centermass(self,
                                flow_threshold: float = None
                                ) -> tuple[dict, dict]:
        """Compute the distance between the initial position of the center of mass and the final furthest point
        for each model loaded using :meth:`load_numerical_result`.

        Parameters
        ----------
        flow_threshold : float, optional
            Flow threshold to extract flow area, if None use 1% of initial maximal 
            flow height. By default None.

        Returns
        -------
        tuple[dict, dict]
            init_coord_centermass : dict
                Coordinates of the initial center of mass for each model: 
                init_coord_centermass[model] = [X, Y, Z].
            model_max_dist : dict
                Maximal distance for each model : model_max_dist[model] = dist.
        """
        init_coord_centermass = {}
        for model in self._loaded_results:
            centermass = self._loaded_results[model].get_output("centermass")
            init_coord_centermass[model] = [centermass.d[0, 0], 
                                            centermass.d[1, 0],
                                            centermass.d[2, 0]]
        
        area_surf, _ = self.compute_impacted_area(flow_threshold=flow_threshold)

        model_max_dist = {}
        for model in self._loaded_results:
            model_max_dist[model] = [None, None, 0.0]
        
        for model in self._loaded_results:
            for i in range(len(area_surf[model].x)):
                for j in range(len(area_surf[model].y)):
                    if area_surf[model].d[j, i] == 1:
                        distance = np.sqrt((area_surf[model].x[i] - init_coord_centermass[model][0])**2 +
                                           (area_surf[model].y[j] - init_coord_centermass[model][1])**2 +
                                           (area_surf[model].z[j,i] - init_coord_centermass[model][2])**2)
                        if distance > model_max_dist[model][2]:
                            model_max_dist[model][0] = area_surf[model].x[i]
                            model_max_dist[model][1] = area_surf[model].y[j]
                            model_max_dist[model][2] = distance
        
        return init_coord_centermass, model_max_dist
    
    
    def compute_average_velocity(self,
                                 distance: float = None,
                                 look_up_direction: str = "right",
                                 flow_threshold: float = None,
                                 **extration_profile_params
                                 ) -> tuple[dict, dict, float, dict]:
        """Compute average velocity for each model loaded using :meth:`load_numerical_result`.

        Parameters
        ----------
        distance : float, optional
            Distance used to calculate average speed, if None use :data:`maximal_distance/2`. 
            By default None.
        look_up_direction : str, optional
            Direction to look for the flow front, must be "right" or "left", 
            by default "right".
        flow_threshold : float, optional
            Flow threshold when extracting front position from profile, if None use 
            1% of initial maximal flow height. By default None.
        
        Returns
        -------
        tuple[dict, dict, float, dict]
            model_avrg_vel: dict
                Average velocity value for each model: model_avrg_vel[model] = avrg_velocity.
            model_time: dict
                Time for each model to reach the distance: model_time[model] = time.
            distance: float
                Distance used to compute the average velocity.
            model_pos: dict
                Front position of the profile of each model: model_pos[model] = [init_front_pos, final_front_pos, 
                maximal_distance].
        """
        def get_front_index(profile):
            idx = np.where(profile <= flow_threshold)[0]
            return (idx[0] - 1) if len(idx) else len(profile) - 1

        def get_back_index(profile):
            idx = np.where(profile <= flow_threshold)[0]
            return (idx[-1] + 1) if len(idx) else 0
        
        extration_profile_params = {} if extration_profile_params is None else extration_profile_params
        
        if look_up_direction not in ["right", "left"]:
            raise ValueError("Invalid look-up direction: 'right' or 'left'")
        
        # Extract profile
        model_profile = {}
        for model in self._loaded_results:
            profile, _ = self._loaded_results[model].get_profile(output="h",
                                                                 **extration_profile_params)
            model_profile[model] = profile
            
            if flow_threshold is None:
                flow_threshold = np.max(profile.d[:, 0])*0.01
            
        # Find initial and final position
        model_pos = {}
        for model in self._loaded_results:
            max_index_init = np.argmax(model_profile[model].d[:, 0])
            max_index_final = np.argmax(model_profile[model].d[:, -1])

            idx_r_init = get_front_index(model_profile[model].d[max_index_init:, 0]) + max_index_init
            idx_l_init = get_back_index(model_profile[model].d[:max_index_init, 0])
            idx_r_final = get_front_index(model_profile[model].d[max_index_final:, -1]) + max_index_final
            idx_l_final = get_back_index(model_profile[model].d[:max_index_final, -1])
            
            if look_up_direction == "right":
                init_pos = model_profile[model].coords[idx_r_init]
                final_pos = model_profile[model].coords[idx_r_final]
            else:
                init_pos = model_profile[model].coords[idx_l_init]
                final_pos = model_profile[model].coords[idx_l_final]

            model_pos[model] = (init_pos, final_pos, abs(final_pos-init_pos))

        # Find minimal distance
        min_distance = 1e10
        for model in model_pos:
            min_distance = model_pos[model][2] if model_pos[model][2] < min_distance else min_distance
        
        if distance is not None and distance > min_distance:
            raise ValueError(f"The requested distance is greater than the minimum distance: {min_distance}")
        
        if distance is None:
            distance = min_distance/2
        
        
        # Find the time for each model to exceed the distance
        model_time = {}
        for model in model_profile:
            for t in range(1, len(model_profile[model].t)):
                max_index = np.argmax(model_profile[model].d[:, t])
                
                idx_r = get_front_index(model_profile[model].d[max_index:, t]) + max_index
                idx_l = get_back_index(model_profile[model].d[:max_index, t])
                
                if look_up_direction == "right":
                    pos = model_profile[model].coords[idx_r]
                else:
                    pos = model_profile[model].coords[idx_l]
                
                if abs(pos - model_pos[model][0]) >= distance:
                    model_time[model] = model_profile[model].t[t]
                    break
        
        
        # Compute the avrg velocity for each model
        model_avrg_vel = {}
        for model in model_time:
            model_avrg_vel[model] = distance / model_time[model]
                    
        return model_avrg_vel, model_time, distance, model_pos


    def compute_rms_from_coussot(self,
                                 coussot_params: dict,
                                 look_up_direction: str = "right",
                                 flow_threshold: float = None,
                                 **extration_profile_params
                                 ) -> tuple[dict, dict, dict]:
        """Compute RMS with Coussot's front shape for each model loaded using :meth:`load_numerical_result`.

        Parameters
        ----------
        coussot_params : dict
            Arguments for generating Coussot's solution. See :class:`tilupy.analytic_sol.Coussot_shape`.
        look_up_direction : str, optional
            Direction to look for the flow front, must be "right" or "left", 
            by default "right".
        flow_threshold : float, optional
            Flow threshold when extracting front position from profile, if None use 
            1% of initial maximal flow height. By default None.

        Returns
        -------
        tuple[dict, dict, dict]
            output_rms: dict
                RMS value for each model: output_rms[model] = rms_value.
            model_front_pos: dict
                Position of Coussot's profile for each model: model_front_pos[model] = pos_value.
            model_coussot: dict
                Coussot's profile for each model: model_coussot[model] = StaticResults1D.
        """        
        def get_front_index(profile):
            idx = np.where(profile <= flow_threshold)[0]
            return (idx[0] - 1) if len(idx) else len(profile) - 1

        def get_back_index(profile):
            idx = np.where(profile <= flow_threshold)[0]
            return (idx[-1] + 1) if len(idx) else 0
        
        output_rms = {}
        model_front_pos = {}
        model_coussot = {}
        for model in self._loaded_results:
            # Extract profile
            prof, _ = self._loaded_results[model].get_profile(output="h",
                                                              **extration_profile_params)
            if flow_threshold is None:
                flow_threshold = np.max(prof.d[:, 0])*0.01
            max_index = np.argmax(prof.d[:, -1])
            # Create Coussot profile
            front_shape = Coussot_shape(**coussot_params, h_final=np.max(prof.d[:, -1]))
            front_shape.compute_rheological_test_front_morpho()
            
            if look_up_direction == "right":
                front_shape.change_orientation_flow()
                idx = get_front_index(prof.d[max_index:, -1]) + max_index
                front_pos = prof.coords[idx]
            else:
                idx = get_back_index(prof.d[:max_index, -1])
                front_pos = prof.coords[idx]                
            front_shape.translate_front(front_pos)
            front_shape.interpolate_on_d()
            
            coussot_pos = front_shape.d.copy()
            
            # Find best position for Coussot's profile
            max_dx = (prof.coords[1] - prof.coords[0]) * 5
            max_index = np.argmax(prof.d[:, -1])
            
            dx_range = np.linspace(-max_dx, max_dx, 5)
            best_dx = 0.0
            min_rms = 1e30
            # x_index = [None, None]
            
            for dx in dx_range:
                temp_pos = coussot_pos + dx
            
                # Extract front value of profile 
                if look_up_direction == "right":
                    x_index_max = np.argmin(np.abs(prof.coords[max_index:] - np.max(temp_pos))) + max_index
                    x_index_min = np.argmin(np.abs(prof.coords[max_index:] - np.min(temp_pos))) + max_index
                else:
                    x_index_max = np.argmin(np.abs(prof.coords[:max_index] - np.max(temp_pos)))
                    x_index_min = np.argmin(np.abs(prof.coords[:max_index] - np.min(temp_pos)))

                reduc_idx = np.linspace(0, len(temp_pos) - 1, len(prof.coords[x_index_min:x_index_max]), dtype=int)
                
                rms = np.sqrt(np.mean(((prof.d[x_index_min:x_index_max, -1]) 
                                    - (front_shape.h[reduc_idx]))**2))
                
                rms /= np.sqrt(np.mean(((front_shape.h[reduc_idx]))**2))

                if rms <= min_rms:
                    min_rms = rms
                    best_dx = dx
                    # x_index = [x_index_min, x_index_max]
            
            output_rms[model] = min_rms
            front_shape.translate_front(best_dx)
            model_front_pos[model] = front_pos + best_dx
            
            model_coussot[model] = tilupy.read.StaticResults1D(name="h",
                                                               d=front_shape.h,
                                                               coords=front_shape.d,
                                                               coords_name="x")
                        
            # prof.plot()
            # plt.plot(front_shape.d[reduc_idx], front_shape.h[reduc_idx], color='red')
            # plt.show()
            
            # plt.plot([0, np.max(front_shape.h[reduc_idx])], [0, np.max(front_shape.h[reduc_idx])], color='red')
            # plt.plot(prof.d[x_index[0]:x_index[1], -1], front_shape.h[reduc_idx])
            # plt.xlabel("Solution numérique")
            # plt.ylabel("Solution de Coussot")
            # plt.show()
        
        return output_rms, model_front_pos, model_coussot

    
    def generate_simulation_comparison_csv(self,
                                           save: bool = False,
                                           folder_out: str = None,
                                           file_name: str = None,
                                           fmt: str = "csv",
                                           available_profile: bool = False,
                                           extration_profile_params: dict = None,
                                           flow_threshold: float = 1e-1,
                                           profile_direction: str = "right",
                                           avrg_velocity_distance: float = None,
                                           coussot_criteria: dict = None,
                                           ) -> None:
        """Generate a csv file summarizing comparison criteria between flow models.
        
        Generates a file grouping comparison criteria between numerical flow models:
        
            - Criteria integrated throughout the simulation:

                - Flow Area: flow area value and difference with mean result (:meth:`compute_impacted_area`).
                - Impacted Zone: RMS of impacted area versus mean result (:meth:`compute_impacted_area_rms_from_avrg`).
                - Maximal Height: RMS of flow maximal height versus mean result (:meth:`compute_rms_from_avrg`).
                - Maximal Momentum: RMS of flow maximal mementum versus mean result (:meth:`compute_rms_from_avrg`).
                - Average Velocity (if profile): time for flow to reach a given distance 
                and average velocity calculated from these values (:meth:`compute_average_velocity`).
            
            - Criteria for the final time step of the simulation:

                - Final Height: RMS of flow final height versus mean result (:meth:`compute_rms_from_avrg`).
                - Maximal Extension: distance from the initial center of mass and the final furthest point of the flow (:meth:`compute_dist_centermass`).
                - Flow Front Position (if profile): maximum distance traveled by the flow and comparison with the average result (:meth:`compute_average_velocity`).
                - Front Shape (optional): RMS of the front shape versus Coussot's theorical front shape (:meth:`compute_rms_from_coussot`).
                
            - Numerical criteria:

                - Volume: value of the volume at final time steps (compared to initial volume) and RMS versus initial volume value.
            
        Parameters
        ----------
        save : bool, optional
            If True, save the resulting tab at :data:`folder_out`. By default False.
        folder_out : str, optional
            Path to the folder where the file is saved, if None generate "xlsx_results"
            folder in code folder. By default None
        file_name : str, optional
            Name of the folder, if None use "results_[models]". By default None.
        fmt : str, optional
            Saving format of the table. Can be "csv" or "xlsx". By default "csv".
        available_profile: bool, optional
            If True, calculate criteria requiring a profile: avrg velocity and final front 
            position. By default False.
        flow_threshold : float, optional
            Flow threshold when extracting front position from profile, by default 1e-1.
        profile_direction : str, optional
            Direction to look for the flow front, must be "right" or "left", 
            by default "right".
        avrg_velocity_distance : float, optional
            Distance used to calculate average speed, if None use :data:`maximal_distance/2`. 
            By default None.
        coussot_criteria : dict, optional
            If None, ignore Coussot criteria. Otherwise, list of arguments to generate a 
            Coussot profile, by default None

        Raises
        ------
        ValueError
            If the file already exists.
        """
        import pandas as pd
        import openpyxl
        import os

        if save:
            if folder_out is None:
                folder_out = os.path.join(os.path.dirname(__file__), "xlsx_results")
                
            os.makedirs(folder_out, exist_ok=True)
            
            if file_name is None:
                file_name = "results"
                for model in self._loaded_results:
                    file_name += "_" + model
            
            if not file_name.endswith("." + fmt):
                file_name = file_name + "." + fmt
            saving_path = os.path.join(folder_out, file_name)
        
            if os.path.exists(saving_path):
                raise ValueError(f"Existing file: {saving_path}")
        
        # Create table columns
        cols = ["", ""]
        for model in self._loaded_results:
            cols.append(model)
        
        table_content = []
        
        if extration_profile_params is None and available_profile:
            extration_profile_params = {} 
        
        # --------------------------------------------------------------------------------------------
        #                        Criteria integrated throughout the simulation 
        # --------------------------------------------------------------------------------------------
        table_content.append(["Criteria integrated throughout the simulation"])
        
        # --------------------------------------- Flow area ------------------------------------------
        line = ["Flow Area", "Total Area [m2]"]
        _, area_num = self.compute_impacted_area(flow_threshold=flow_threshold)
        
        mean_list = []
        for model in self._loaded_results:
            mean_list.append(area_num[model])
            line.append(area_num[model])
        table_content.append(line)
        
        line = ["", "Compared to avrg"]
        mean_area = np.mean(np.array(mean_list))
        for model in self._loaded_results:
            line.append((area_num[model] - mean_area)/mean_area)
        table_content.append(line)    
            
        line = ["Impacted Zone", "RMS (avrg)"]
        rms_area, _ = self.compute_impacted_area_rms_from_avrg(flow_threshold=flow_threshold)
        
        for model in self._loaded_results:
            line.append(rms_area[model])
        table_content.append(line)

        # ------------------------------------- Maximal height ---------------------------------------
        line = ["Maximal Height", "RMS (avrg)"]
        rms_height = self.compute_rms_from_avrg(output="h_max")

        for model in self._loaded_results:
            line.append(rms_height[model])
        table_content.append(line)
        
        # ------------------------------------- Maximal momentum ---------------------------------------
        line = ["Maximal Momentum", "RMS (avrg)"]
        rms_hu = self.compute_rms_from_avrg(output="hu_max")

        for model in self._loaded_results:
            line.append(rms_hu[model])
        table_content.append(line)

        # ----------------------------------- Average flow speed -------------------------------------
        if available_profile:
            avrg_vel, time, dist, model_pos = self.compute_average_velocity(distance=avrg_velocity_distance,
                                                                            look_up_direction=profile_direction,
                                                                            flow_threshold=flow_threshold,
                                                                            **extration_profile_params)
            line = ["Velocity", f"Time [s] to complete d={dist}m"]

            for model in self._loaded_results:
                line.append(time[model])
            table_content.append(line)
            
            line = ["", "Average velocity [m/s]"]
            
            for model in self._loaded_results:
                line.append(avrg_vel[model])
            table_content.append(line)


        # --------------------------------------------------------------------------------------------
        #                       Criteria for the final time step of the simulation
        # --------------------------------------------------------------------------------------------
        table_content.append(["Criteria for the final time step of the simulation"])

        # --------------------------------------- Final height ---------------------------------------
        line = ["Final Thickness Repartition", "RMS (avrg)"]
        rms_height = self.compute_rms_from_avrg(output="h_final")

        for model in self._loaded_results:
            line.append(rms_height[model])
        table_content.append(line)
        
        # -------------------------------- Position of the flow front --------------------------------
        line = ["Maximal Extension", "Distance [m]"]
        _, model_max_dist = self.compute_dist_centermass(flow_threshold=flow_threshold)
        
        for model in self._loaded_results:
            line.append(model_max_dist[model][2])
        table_content.append(line)
        
        line = ["", "Compared to avrg"]

        list_pos = []
        for model in self._loaded_results:
            list_pos.append(model_max_dist[model][2])
        mean_dist = np.mean(np.array(list_pos))
        
        for model in self._loaded_results:
            line.append((model_max_dist[model][2]-mean_dist)/mean_dist)
        table_content.append(line)
        
        if available_profile:
            line = ["Flow Front (Profile)", "Distance [m]"]

            for model in self._loaded_results:
                line.append(model_pos[model][2])
            table_content.append(line)
            
            line = ["", "Compared to avrg"]

            list_pos = []
            for model in self._loaded_results:
                list_pos.append(model_pos[model][2])
            mean_dist = np.mean(np.array(list_pos))
            
            for model in self._loaded_results:
                line.append((model_pos[model][2]-mean_dist)/mean_dist)
            table_content.append(line)

        # ---------------------------------- Front shape comparison ---------------------------------- 
        if coussot_criteria is not None:
            line = ["Front Shape", "RMS (Coussot)"]
            model_rms, _, _ = self.compute_rms_from_coussot(look_up_direction=profile_direction,
                                                            flow_threshold=flow_threshold,
                                                            coussot_params=coussot_criteria)
            
            for model in self._loaded_results:
                line.append(model_rms[model])
            table_content.append(line)

            
        # --------------------------------------------------------------------------------------------
        #                                     Numeralical criterias
        # --------------------------------------------------------------------------------------------
        table_content.append(["Numerical criteria"])
        
        # ------------------------------------------- Volume -----------------------------------------
        model_volume = {}
        for model in self._loaded_results:
            volume = self._loaded_results[model].get_output("volume")
            model_volume[model] = volume
            
        line = ["Volume Variation", "RMS of Total Volume Variation"]
                        
        for model in self._loaded_results:
            rms_vinit = 0
            for v in model_volume[model].d:
                rms_vinit += np.sqrt((v - model_volume[model].d[0])**2) / model_volume[model].d[0]
            line.append(rms_vinit)
        table_content.append(line)
        
        line = ["", "(Vf - Vi) / Vi"]
        
        for model in self._loaded_results:
            line.append((model_volume[model].d[-1] - model_volume[model].d[0]) / 
                        model_volume[model].d[0])
        table_content.append(line)
        
        
        # --------------------------------------------------------------------------------------------
        #                                          Saving table
        # --------------------------------------------------------------------------------------------
        df = pd.DataFrame(table_content, columns=cols)
        if save:
            if fmt == "csv":
                df.to_csv(saving_path, index=False, encoding="utf-8")
                
            if fmt == "xlsx":
                df.to_excel(saving_path, index=False, engine="openpyxl")
                
        # --------------------------------------- Reformat table -------------------------------------
                wb = openpyxl.load_workbook(saving_path)
                ws = wb.active
                
                # Add border style
                bordure_fine = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                    right=openpyxl.styles.Side(style="thin"),
                                                    top=openpyxl.styles.Side(style="thin"),
                                                    bottom=openpyxl.styles.Side(style="thin"))

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.border = bordure_fine
                wb.save(saving_path)
            
            print(f"Saved in: {saving_path}")
        
        return df


    def generate_analytical_comparison_csv(self,
                                           analytic_solution: dict,
                                           save: bool = False,
                                           folder_out: str = None,
                                           file_name: str = None,
                                           fmt: str = "csv",
                                           compute_as_u: bool = True,
                                           extration_profile_params: dict = None,
                                           flow_threshold: float = None,
                                           ) -> None:
        """Generate a csv file summarizing comparison criteria between flow models and analytic solution.
        
        Generates a file grouping comparison criteria between numerical flow models and analytic solution:
        
            - Criteria integrated throughout the simulation:

                - Total Height: RMS of flow height versus analytical solution.
                - Total Momentum: RMS of flow h*u versus analytical solution.
                - Total Front Position: RMS of flow front versus analytical solution.
            
            - Criteria for the final time step of the simulation:

                - Final Height: RMS of flow final height versus analytical solution.
                - Flow Front Position: maximum distance traveled by the flow and comparison with the analytic solution.
                 
            - Numerical criteria:

                - Volume: value of the volume at final time steps (compared to initial volume) and RMS versus initial volume value.
            
        Parameters
        ----------
        analytic_solution: dict
            Argument for the analytic solution. See :meth:`compute_analytic_solution`.
        save : bool, optional
            If True, save the resulting tab at :data:`folder_out`. By default False.
        folder_out : str, optional
            Path to the folder where the file is saved, if None generate "xlsx_results"
            folder in code folder. By default None
        file_name : str, optional
            Name of the folder, if None use "AS_comparison_[models]". By default None.
        fmt : str, optional
            Saving format of the table. Can be "csv" or "xlsx". By default "csv".
        compute_as_u: bool, optional
            If True, compute analytic solution for flow velocity. Can be disabled. By default True.
        extration_profile_params: dict, optional
            Argument for extracting profile. See :meth:`tilupy.read.Results.get_profile`. By default None.
        flow_threshold : float, optional
            Flow threshold when extracting front position from profile, by default None.
        
        Raises
        ------
        ValueError
            If the file already exists.
        """
        import pandas as pd
        import openpyxl
        import os

        if save:
            if folder_out is None:
                folder_out = os.path.join(os.path.dirname(__file__), "xlsx_results")
                
            os.makedirs(folder_out, exist_ok=True)
            
            if file_name is None:
                file_name = "AS_compare"
                for model in self._loaded_results:
                    file_name += "_" + model
            
            if not file_name.endswith("." + fmt):
                file_name = file_name + "." + fmt
            saving_path = os.path.join(folder_out, file_name)
        
            if os.path.exists(saving_path):
                raise ValueError(f"Existing file: {saving_path}")
        
        extration_profile_params = {} if extration_profile_params is None else extration_profile_params
        
        # Create table columns
        cols = ["", ""]
        for model in self._loaded_results:
            cols.append(model)
        
        table_content = []
        
        # --------------------------------------------------------------------------------------------
        #                        Compute AS at each time steps for each models 
        # --------------------------------------------------------------------------------------------
        model_AS_h = {}
        model_AS_u = {}
        for model in self._loaded_results:
            as_h_profile = self.compute_analytic_solution(output="h",
                                                          T=self._models_tim[model],
                                                          model=model,
                                                          **analytic_solution)
            model_AS_h[model] = as_h_profile
            if compute_as_u:
                as_u_profile = self.compute_analytic_solution(output="u",
                                                              T=self._models_tim[model],
                                                              model=model,
                                                              **analytic_solution)
                
                model_AS_u[model] = np.nan_to_num(as_u_profile.d, nan=0)
        
            if flow_threshold is None:
                flow_threshold = np.max(model_AS_h[model].d) * 0.01
        
        # --------------------------------------------------------------------------------------------
        #                        Criteria integrated throughout the simulation 
        # --------------------------------------------------------------------------------------------
        table_content.append(["Criteria integrated throughout the simulation"])
        
        # ---------------------------------- Height difference ---------------------------------------
        model_h = {}
        for model in self._loaded_results:
            profile, _ = self._loaded_results[model].get_profile(output="h",
                                                                 **extration_profile_params)
            model_h[model] = profile
            
        line = ["Total Height Difference", "RMS (AS)"]
        for model in self._loaded_results:
            line.append((np.sqrt(np.sum((model_h[model].d - model_AS_h[model].d)**2)) /
                         np.sqrt(np.sum((model_AS_h[model].d)**2))))
        
        table_content.append(line)
        
        # ------------------------------------ HU difference -----------------------------------------
        if compute_as_u:
            model_u = {}
            for model in self._loaded_results:
                profile, _ = self._loaded_results[model].get_profile(output="u",
                                                                     **extration_profile_params)
                model_u[model] = profile
                
            line = ["Total Momentum Difference", "RMS (AS)"]
            for model in self._loaded_results:
                hu = model_h[model].d * model_u[model].d
                as_hu = model_AS_h[model].d * model_AS_u[model]
                line.append((np.sqrt(np.sum((hu - as_hu)**2)) /
                            np.sqrt(np.sum((as_hu)**2))))
            
            table_content.append(line)
        
        # ----------------------------------- Front position ---------------------------------------
        model_position = {}
        model_as_position = {}
        for model in self._loaded_results:
            list_position = []
            list_as_position = []
            for t in range(len(model_h[model].t)):
                # Model
                max_index = np.argmax(model_h[model].d[:, t])
                idx = np.where(model_h[model].d[max_index:, t] <= flow_threshold)[0]
                idx = (idx[0] - 1) if len(idx) else len(model_h[model].d[max_index:, t]) - 1
                idx += max_index
                list_position.append(model_h[model].coords[idx])
                # AS
                max_index = np.argmax(model_AS_h[model].d[:, t])
                idx = np.where(model_AS_h[model].d[max_index:, t] <= flow_threshold)[0]
                idx = (idx[0] - 1) if len(idx) else len(model_AS_h[model].d[max_index:, t]) - 1
                idx += max_index
                list_as_position.append(model_AS_h[model].coords[idx])
            model_position[model] = np.array(list_position)
            model_as_position[model] = np.array(list_as_position) 
        
        line = ["Total Front Position Difference", "RMS (AS)"]
        
        for model in self._loaded_results:
            line.append((np.sqrt(np.sum((model_position[model] - model_as_position[model])**2)) /
                        np.sqrt(np.sum((model_as_position[model])**2))))
        
        table_content.append(line)
        
        # --------------------------------------------------------------------------------------------
        #                       Criteria for the final time step of the simulation
        # --------------------------------------------------------------------------------------------
        table_content.append(["Criteria for the final time step of the simulation"])
        
        # --------------------------------------- Final height ---------------------------------------
        line = ["Final Thickness Repartition", "RMS (AS)"]
        for model in self._loaded_results:
            line.append((np.sqrt(np.sum((model_h[model].d[:, -1] - model_AS_h[model].d[:, -1])**2)) /
                         np.sqrt(np.sum((model_AS_h[model].d[:, -1])**2))))
        
        table_content.append(line)
        
        # ---------------------------- Front position at final time step -----------------------------
        line = ["Final Front Position", "Distance [m]"]
        
        for model in self._loaded_results:
            line.append(model_position[model][-1] - model_position[model][0])
        
        table_content.append(line)
        
        line = ["", "Compared to AS"]
        
        # model_as_position[model] différent selon model
        for model in self._loaded_results:
            line.append((model_position[model][-1] - model_as_position[model][-1]) /
                         model_as_position[model][-1])
        
        table_content.append(line)
        
        # --------------------------------------------------------------------------------------------
        #                                     Numeralical criteria
        # --------------------------------------------------------------------------------------------
        table_content.append(["Numerical criteria"])
        
        # ------------------------------------------- Volume -----------------------------------------
        model_volume = {}
        for model in self._loaded_results:
            volume = self._loaded_results[model].get_output("volume")
            model_volume[model] = volume
            
        line = ["Volume Variation", "RMS of Total Volume Variation"]
                        
        for model in self._loaded_results:
            rms_vinit = 0
            for v in model_volume[model].d:
                rms_vinit += np.sqrt((v - model_volume[model].d[0])**2) / model_volume[model].d[0]
            line.append(rms_vinit)
        table_content.append(line)
        
        line = ["", "(Vf - Vi) / Vi"]
        
        for model in self._loaded_results:
            line.append((model_volume[model].d[-1] - model_volume[model].d[0]) / 
                        model_volume[model].d[0])
        table_content.append(line)
            
            
        # --------------------------------------------------------------------------------------------
        #                                          Saving table
        # --------------------------------------------------------------------------------------------
        df = pd.DataFrame(table_content, columns=cols)
        if save:
            if fmt == "csv":
                df.to_csv(saving_path, index=False, encoding="utf-8")
                
            if fmt == "xlsx":
                df.to_excel(saving_path, index=False, engine="openpyxl")
                
        # --------------------------------------- Reformat table -------------------------------------
                wb = openpyxl.load_workbook(saving_path)
                ws = wb.active
                
                # Add border style
                bordure_fine = openpyxl.styles.Border(left=openpyxl.styles.Side(style="thin"),
                                                    right=openpyxl.styles.Side(style="thin"),
                                                    top=openpyxl.styles.Side(style="thin"),
                                                    bottom=openpyxl.styles.Side(style="thin"))

                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.border = bordure_fine
                wb.save(saving_path)
            
            print(f"Saved in: {saving_path}")
        
        return df
        